from backend.database.connection import attendance_collection, event_collection, user_collection
from backend.models.attendance_model import Attendance
from backend.utils.location_validator import is_within_radius
from backend.utils.serializer import serialize_doc
from datetime import datetime
from bson import ObjectId
import face_recognition
import numpy as np
from fastapi.concurrency import run_in_threadpool
import io

def _verify_face_sync(known_encoding, image_bytes):
    try:
        unknown_image = face_recognition.load_image_file(io.BytesIO(image_bytes))
        unknown_encodings = face_recognition.face_encodings(unknown_image)
        
        if not unknown_encodings:
            return False
            
        unknown_encoding = unknown_encodings[0]
        results = face_recognition.compare_faces([np.array(known_encoding)], unknown_encoding, tolerance=0.6)
        return bool(results[0])
    except Exception as e:
        print(f"Face verification error: {e}")
        return False

async def check_in(student_id: str, event_id: ObjectId, user_lat: float, user_lon: float, face_image_bytes: bytes = None):
    from geopy.distance import geodesic
    
    event = await event_collection.find_one({"_id": event_id})
    if not event:
        return None, "Event not found"

    # Check if check-in is currently allowed
    # Use local time - no UTC conversion
    now = datetime.now()
    check_in_start_raw = event.get("check_in_start") or event.get("start_time")
    check_in_end_raw = event.get("check_in_end") or event.get("end_time")
    
    # MongoDB stores datetimes as UTC. We need to treat them as naive local times.
    # Strip any timezone information and treat as local time
    check_in_start = None
    check_in_end = None
    
    if check_in_start_raw:
        if isinstance(check_in_start_raw, datetime):
            # If it has timezone info, remove it (treat the time as-is in local timezone)
            check_in_start = check_in_start_raw.replace(tzinfo=None) if check_in_start_raw.tzinfo else check_in_start_raw
        else:
            check_in_start = check_in_start_raw
    
    if check_in_end_raw:
        if isinstance(check_in_end_raw, datetime):
            check_in_end = check_in_end_raw.replace(tzinfo=None) if check_in_end_raw.tzinfo else check_in_end_raw
        else:
            check_in_end = check_in_end_raw
    
    print(f"[CHECK-IN] Time validation:")
    print(f"   Current time (local): {now.strftime('%Y-%m-%d %I:%M:%S %p')}")
    print(f"   Check-in start (stored): {check_in_start.strftime('%Y-%m-%d %I:%M:%S %p') if check_in_start else 'N/A'}")
    print(f"   Check-in end (stored): {check_in_end.strftime('%Y-%m-%d %I:%M:%S %p') if check_in_end else 'N/A'}")
    print(f"   Comparison: now={now}, check_in_start={check_in_start}")
    print(f"   Can check in? {not (check_in_start and now < check_in_start)}")
    
    if check_in_start and now < check_in_start:
        return None, f"Check-in not yet open. Opens at {check_in_start.strftime('%I:%M %p')}"
    
    if check_in_end and now > check_in_end:
        return None, "Check-in period has ended"
    
    if check_in_end and now > check_in_end:
        return None, "Check-in period has ended"

    # Check if student already checked in for this event
    existing_attendance = await attendance_collection.find_one({
        "student_id": student_id,
        "event_id": event_id
    })
    
    if existing_attendance and existing_attendance.get("check_in_time"):
        print(f"[WARNING] Student {student_id} already checked in for event {event.get('name')}")
        return None, "Already checked in"

    # Calculate distance for debugging
    event_location = (event["latitude"], event["longitude"])
    user_location = (user_lat, user_lon)
    distance = geodesic(user_location, event_location).meters
    
    print(f"[DEBUG] Check-in Debug:")
    print(f"   Event: {event.get('name', 'Unknown')}")
    print(f"   Event Location: {event['latitude']}, {event['longitude']}")
    print(f"   User Location: {user_lat}, {user_lon}")
    print(f"   Distance: {distance:.2f} meters")
    print(f"   Allowed Radius: {event['radius']} meters")
    
    # Only record attendance if student is within range (Present)
    if is_within_radius(user_lat, user_lon, event["latitude"], event["longitude"], event["radius"]):
        
        # Face Verification Logic
        if face_image_bytes:
            user = await user_collection.find_one({"_id": ObjectId(student_id)})
            if user:
                if user.get("face_encoding"):
                    is_match = await run_in_threadpool(_verify_face_sync, user["face_encoding"], face_image_bytes)
                    if not is_match:
                        return None, "Face verification failed. Face does not match registered user."
                else:
                    return None, "Face not registered. Please register your face in profile settings."
            else:
                return None, "User not found"
        
        print(f"   [OK] Status: PRESENT (within range)")
        
        # Create or update attendance record
        attendance_dict = {
            "student_id": student_id,
            "event_id": event_id,
            "check_in_time": now,
            "check_in_status": "Present",
            "timestamp": now,
            "status": "Present"
        }

        if existing_attendance:
            # Update existing record with check-in
            await attendance_collection.update_one(
                {"_id": existing_attendance["_id"]},
                {"$set": attendance_dict}
            )
            updated_attendance = await attendance_collection.find_one({"_id": existing_attendance["_id"]})
            return serialize_doc(updated_attendance), "Checked in successfully"
        else:
            # Create new attendance record
            new_attendance = await attendance_collection.insert_one(attendance_dict)
            created_attendance = await attendance_collection.find_one({"_id": new_attendance.inserted_id})
            return serialize_doc(created_attendance), "Checked in successfully"
    else:
        print(f"   [ERROR] Status: OUT OF RANGE (distance exceeds radius)")
        return None, "Out of Range"

async def check_out(student_id: str, event_id: ObjectId, user_lat: float, user_lon: float, face_image_bytes: bytes = None):
    from geopy.distance import geodesic
    
    event = await event_collection.find_one({"_id": event_id})
    if not event:
        return None, "Event not found"

    # Check if check-out is currently allowed
    # Use local time - no UTC conversion
    now = datetime.now()
    check_out_start_raw = event.get("check_out_start") or event.get("end_time")
    check_out_end_raw = event.get("check_out_end")
    
    # MongoDB stores datetimes as UTC. Strip timezone info and treat as local time
    check_out_start = None
    check_out_end = None
    
    if check_out_start_raw:
        if isinstance(check_out_start_raw, datetime):
            check_out_start = check_out_start_raw.replace(tzinfo=None) if check_out_start_raw.tzinfo else check_out_start_raw
        else:
            check_out_start = check_out_start_raw
    
    if check_out_end_raw:
        if isinstance(check_out_end_raw, datetime):
            check_out_end = check_out_end_raw.replace(tzinfo=None) if check_out_end_raw.tzinfo else check_out_end_raw
        else:
            check_out_end = check_out_end_raw
    
    print(f"[CHECK-OUT] Time validation:")
    print(f"   Current time (local): {now.strftime('%I:%M %p')}")
    print(f"   Check-out start (stored): {check_out_start.strftime('%I:%M %p') if check_out_start else 'N/A'}")
    
    if check_out_start and now < check_out_start:
        return None, f"Check-out not yet open. Opens at {check_out_start.strftime('%I:%M %p')}"
    
    if check_out_end and now > check_out_end:
        return None, "Check-out period has ended"
    
    if check_out_end and now > check_out_end:
        return None, "Check-out period has ended"

    # Check if student has checked in
    existing_attendance = await attendance_collection.find_one({
        "student_id": student_id,
        "event_id": event_id
    })
    
    if not existing_attendance or not existing_attendance.get("check_in_time"):
        return None, "You must check in first before checking out"
    
    if existing_attendance.get("check_out_time"):
        return None, "Already checked out"

    # Calculate distance
    event_location = (event["latitude"], event["longitude"])
    user_location = (user_lat, user_lon)
    distance = geodesic(user_location, event_location).meters
    
    print(f"[DEBUG] Check-out Debug:")
    print(f"   Event: {event.get('name', 'Unknown')}")
    print(f"   Distance: {distance:.2f} meters")
    
    # Record check-out if within range
    if is_within_radius(user_lat, user_lon, event["latitude"], event["longitude"], event["radius"]):
        
        # Face Verification Logic
        if face_image_bytes:
            user = await user_collection.find_one({"_id": ObjectId(student_id)})
            if user:
                if user.get("face_encoding"):
                    is_match = await run_in_threadpool(_verify_face_sync, user["face_encoding"], face_image_bytes)
                    if not is_match:
                        return None, "Face verification failed. Face does not match registered user."
                else:
                    return None, "Face not registered. Please register your face in profile settings."
            else:
                return None, "User not found"
        
        print(f"   [OK] Check-out: PRESENT (within range)")
        
        # Update attendance with check-out
        await attendance_collection.update_one(
            {"_id": existing_attendance["_id"]},
            {"$set": {
                "check_out_time": now,
                "check_out_status": "Present"
            }}
        )
        
        updated_attendance = await attendance_collection.find_one({"_id": existing_attendance["_id"]})
        return serialize_doc(updated_attendance), "Checked out successfully"
    else:
        print(f"   [ERROR] Check-out: OUT OF RANGE")
        return None, "Out of Range"

async def get_attendance_history(student_id: str):
    """Get attendance history for a student with event names and society/department"""
    from backend.database.connection import department_collection
    
    history = []
    async for record in attendance_collection.find({"student_id": student_id}):
        # Get event details to include event name and department
        event_id = record.get("event_id")
        event_name = "Unknown Event"
        department_name = "Unknown Society"
        
        if event_id:
            event = await event_collection.find_one({"_id": event_id})
            if event:
                event_name = event.get("name", "Unknown Event")
                
                # Get department/society name
                department_id = event.get("department_id")
                if department_id:
                    if isinstance(department_id, str):
                        department_id = ObjectId(department_id)
                    department = await department_collection.find_one({"_id": department_id})
                    if department:
                        department_name = department.get("name", "Unknown Society")
        
        record_data = serialize_doc(record)
        record_data["event_name"] = event_name
        record_data["department_name"] = department_name
        history.append(record_data)
    
    return history

async def get_attendance_status(student_id: str, event_id: ObjectId):
    """Check if a student has checked in/out for an event"""
    attendance = await attendance_collection.find_one({
        "student_id": student_id,
        "event_id": event_id
    })
    
    if not attendance:
        return {
            "has_checked_in": False,
            "has_checked_out": False
        }
    
    return {
        "has_checked_in": bool(attendance.get("check_in_time")),
        "has_checked_out": bool(attendance.get("check_out_time"))
    }

async def get_event_attendance(event_id: ObjectId):
    """Get all attendance records for a specific event with student details
    Includes all enrolled students (Present or Absent)"""
    from backend.database.connection import user_collection, enrollment_collection
    
    # Get the event
    event = await event_collection.find_one({"_id": event_id})
    if not event:
        return []
    
    department_id = event.get("department_id")
    
    # Get all enrolled students in the department
    enrolled_students = {}
    if department_id:
        async for enrollment in enrollment_collection.find({
            "department_id": department_id,
            "status": "approved"
        }):
            student_id = str(enrollment["user_id"])
            enrolled_students[student_id] = {
                "student_id": student_id,
                "status": "Absent",  # Default to Absent
                "timestamp": None
            }
    
    # Update with actual attendance records
    async for record in attendance_collection.find({"event_id": event_id}):
        student_id = record["student_id"]
        if student_id in enrolled_students or not department_id:
            enrolled_students[student_id] = {
                "student_id": student_id,
                "status": record.get("status", "Absent"),
                "timestamp": record.get("timestamp"),
                "check_in_time": record.get("check_in_time"),
                "check_out_time": record.get("check_out_time"),
                "_id": str(record.get("_id", ""))
            }
    
    # Build attendance records with student details
    attendance_records = []
    for student_id, attendance_info in enrolled_students.items():
        # Get student information
        student = await user_collection.find_one({"_id": ObjectId(student_id)})
        
        attendance_data = {
            "_id": attendance_info.get("_id", ""),
            "student_id": student_id,
            "status": attendance_info["status"],
            "timestamp": attendance_info["timestamp"],
            "check_in_time": attendance_info.get("check_in_time"),
            "check_out_time": attendance_info.get("check_out_time")
        }
        
        if student:
            # Build name from first_name + last_name if available, otherwise use 'name' field
            first_name = student.get("first_name", "")
            last_name = student.get("last_name", "")
            if first_name and last_name:
                full_name = f"{first_name} {last_name}"
            else:
                full_name = student.get("name", "Unknown")
            
            attendance_data["student_name"] = full_name
            attendance_data["student_email"] = student.get("email", "Unknown")
            attendance_data["student_id_number"] = student.get("id_number", "N/A")
            attendance_data["student_first_name"] = first_name
            attendance_data["student_last_name"] = last_name
            
            # Debug: Log what we're getting from the database
            print(f"[DEBUG] Student {student_id} data:")
            print(f"   Name: {full_name}")
            print(f"   Email: {student.get('email')}")
            print(f"   ID Number: {student.get('id_number', 'N/A')} (field exists: {'id_number' in student})")
            print(f"   First Name: {first_name} (field exists: {'first_name' in student})")
            print(f"   Last Name: {last_name} (field exists: {'last_name' in student})")
            print(f"   ALL STUDENT FIELDS: {list(student.keys())}")
        else:
            attendance_data["student_name"] = "Unknown"
            attendance_data["student_email"] = "Unknown"
            attendance_data["student_id_number"] = "N/A"
            attendance_data["student_first_name"] = ""
            attendance_data["student_last_name"] = ""
        
        attendance_records.append(attendance_data)
    
    return attendance_records

async def export_event_attendance_to_excel(event_id: ObjectId):
    """Export event attendance to Excel format - only Present students with first/last name"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from backend.database.connection import user_collection
    from io import BytesIO
    
    # Get event details
    event = await event_collection.find_one({"_id": event_id})
    if not event:
        raise ValueError("Event not found")
    
    # Get all attendance records (Present and Absent)
    attendance_records = await get_event_attendance(event_id)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    
    # Add header with event info
    ws.merge_cells('A1:E1')
    ws['A1'] = f"Attendance Report - {event.get('name', 'Event')}"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:E2')
    ws['A2'] = f"Date: {event.get('start_time', 'N/A')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Column headers
    headers = ['No.', 'ID Number', 'Last Name', 'First Name', 'Check-In Time', 'Check-Out Time', 'Duration', 'Status']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Helper function to calculate duration
    def calculate_duration(check_in_time, check_out_time):
        if not check_in_time or not check_out_time:
            return '—'
        try:
            delta = check_out_time - check_in_time
            hours = delta.seconds // 3600
            minutes = (delta.seconds % 3600) // 60
            if hours > 0:
                return f"{hours}h {minutes}m"
            else:
                return f"{minutes}m"
        except:
            return 'Invalid'
    
    # Add data - only Present students
    for idx, record in enumerate(attendance_records, 1):
        # Get first name and last name from database, fallback to splitting full name
        first_name = record.get('student_first_name', '')
        last_name = record.get('student_last_name', '')
        id_number = record.get('student_id_number', 'N/A')
        
        # If first_name or last_name is empty, try splitting the full name
        if not first_name or not last_name:
            full_name = record.get('student_name', 'Unknown')
            name_parts = full_name.strip().split(' ', 1)  # Split on first space
            
            if len(name_parts) == 2:
                first_name = name_parts[0] if not first_name else first_name
                last_name = name_parts[1] if not last_name else last_name
            elif len(name_parts) == 1:
                first_name = name_parts[0] if not first_name else first_name
        
        # Get check-in and check-out times
        check_in_time = record.get('check_in_time') or record.get('timestamp')
        check_out_time = record.get('check_out_time')
        
        # Format times
        check_in_str = check_in_time.strftime('%Y-%m-%d %I:%M %p') if check_in_time else 'N/A'
        check_out_str = check_out_time.strftime('%Y-%m-%d %I:%M %p') if check_out_time else 'Not checked out'
        duration_str = calculate_duration(check_in_time, check_out_time)
        
        ws.cell(row=idx+4, column=1, value=idx)
        ws.cell(row=idx+4, column=2, value=id_number)
        ws.cell(row=idx+4, column=3, value=last_name)
        ws.cell(row=idx+4, column=4, value=first_name)
        ws.cell(row=idx+4, column=5, value=check_in_str)
        ws.cell(row=idx+4, column=6, value=check_out_str)
        ws.cell(row=idx+4, column=7, value=duration_str)
        
        status = record.get('status', 'Absent')
        status_cell = ws.cell(row=idx+4, column=8, value=status)
        
        # Color code: Green for Present, Red for Absent
        if status == 'Present':
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        status_cell.alignment = Alignment(horizontal='center')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8   # No.
    ws.column_dimensions['B'].width = 15  # ID Number
    ws.column_dimensions['C'].width = 20  # Last Name
    ws.column_dimensions['D'].width = 20  # First Name
    ws.column_dimensions['E'].width = 22  # Check-In Time
    ws.column_dimensions['F'].width = 22  # Check-Out Time
    ws.column_dimensions['G'].width = 12  # Duration
    ws.column_dimensions['H'].width = 15  # Status
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()

async def finalize_event_attendance(event_id: ObjectId):
    """Mark all enrolled students who didn't check in as Absent"""
    from backend.database.connection import user_collection, enrollment_collection
    
    # Get the event
    event = await event_collection.find_one({"_id": event_id})
    if not event:
        raise ValueError("Event not found")
    
    # Get all students enrolled in the event's department
    department_id = event.get("department_id")
    if not department_id:
        return {"message": "Event has no department", "absent_count": 0}
    
    # Get all enrolled students in this department
    enrolled_students = []
    async for enrollment in enrollment_collection.find({
        "department_id": department_id,
        "status": "approved"
    }):
        enrolled_students.append(str(enrollment["user_id"]))
    
    # Get students who already checked in (Present)
    present_students = []
    async for record in attendance_collection.find({
        "event_id": event_id,
        "status": "Present"
    }):
        present_students.append(record["student_id"])
    
    # Mark absent students
    absent_count = 0
    for student_id in enrolled_students:
        if student_id not in present_students:
            # Check if already marked as absent
            existing = await attendance_collection.find_one({
                "student_id": student_id,
                "event_id": event_id
            })
            
            if not existing:
                # Create absent attendance record
                # Use local time
                attendance_dict = {
                    "student_id": student_id,
                    "event_id": event_id,
                    "timestamp": datetime.now(),
                    "status": "Absent"
                }
                await attendance_collection.insert_one(attendance_dict)
                absent_count += 1
    
    return {
        "message": "Attendance finalized",
        "total_enrolled": len(enrolled_students),
        "present_count": len(present_students),
        "absent_count": absent_count
    }
