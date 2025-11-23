from fastapi import APIRouter, Body, HTTPException, status, Depends
from fastapi.responses import StreamingResponse
from backend.controllers import admin_controller
from backend.models.department_model import Department
from backend.utils.jwt_handler import decodeJWT
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

router = APIRouter(
    prefix="/admin",
    tags=["Admin"],
)

@router.post("/departments", response_description="Add new department")
async def add_department(department: Department = Body(...), token: dict = Depends(decodeJWT)):
    try:
        if not token or token.get("role") != "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Admin access required",
            )
        new_department = await admin_controller.add_department(department)
        
        # Convert ObjectId to string for JSON serialization
        department_dict = dict(new_department)
        department_dict["_id"] = str(department_dict["_id"])
        
        return {
            "message": "Department created successfully",
            "department_id": str(new_department["_id"]),
            "department": department_dict
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to create department: {str(e)}",
        )

@router.get("/departments", response_description="Get all departments")
async def get_departments(token: dict = Depends(decodeJWT)):
    try:
        if not token or token.get("role") != "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Admin access required",
            )
        from backend.database.connection import department_collection
        departments = await department_collection.find().to_list(1000)
        
        # Convert ObjectId to string for each department
        for dept in departments:
            dept["_id"] = str(dept["_id"])
        
        return departments
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch departments: {str(e)}",
        )

@router.put("/departments/{department_id}", response_description="Update department")
async def update_department(department_id: str, name: str = Body(..., embed=True), token: dict = Depends(decodeJWT)):
    try:
        if not token or token.get("role") != "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Admin access required",
            )
        from backend.database.connection import department_collection
        from bson import ObjectId
        
        result = await department_collection.update_one(
            {"_id": ObjectId(department_id)},
            {"$set": {"name": name}}
        )
        
        if result.modified_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Department not found or no changes made",
            )
        
        return {"message": "Department updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update department: {str(e)}",
        )

@router.delete("/departments/{department_id}", response_description="Delete department")
async def delete_department(department_id: str, token: dict = Depends(decodeJWT)):
    try:
        if not token or token.get("role") != "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Admin access required",
            )
        from backend.database.connection import department_collection
        from bson import ObjectId
        
        result = await department_collection.delete_one({"_id": ObjectId(department_id)})
        
        if result.deleted_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Department not found",
            )
        
        return {"message": "Department deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to delete department: {str(e)}",
        )

@router.get("/export/attendance", response_description="Export attendance to Excel")
async def export_attendance(department_id: str = None, token: dict = Depends(decodeJWT)):
    try:
        if not token or token.get("role") != "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Admin access required",
            )
        
        # Get attendance data
        from backend.database.connection import attendance_collection, user_collection, event_collection
        
        query = {}
        if department_id:
            # Get users from this department
            users = await user_collection.find({"department_id": department_id}).to_list(1000)
            user_ids = [str(user["_id"]) for user in users]
            query = {"user_id": {"$in": user_ids}}
        
        attendance_records = await attendance_collection.find(query).to_list(1000)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Add header styling
        header_fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = ["User ID", "User Name", "User Email", "Event ID", "Event Name", "Check-in Time", "Status"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Add data
        row_num = 2
        for record in attendance_records:
            user = await user_collection.find_one({"_id": record["user_id"]})
            event = await event_collection.find_one({"_id": record["event_id"]})
            
            ws.cell(row=row_num, column=1, value=str(record["user_id"]))
            ws.cell(row=row_num, column=2, value=user.get("name", "N/A") if user else "N/A")
            ws.cell(row=row_num, column=3, value=user.get("email", "N/A") if user else "N/A")
            ws.cell(row=row_num, column=4, value=str(record["event_id"]))
            ws.cell(row=row_num, column=5, value=event.get("name", "N/A") if event else "N/A")
            ws.cell(row=row_num, column=6, value=str(record.get("timestamp", "N/A")))
            ws.cell(row=row_num, column=7, value=record.get("status", "N/A"))
            
            row_num += 1
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Return as downloadable file
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=attendance_report.xlsx"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to export attendance: {str(e)}",
        )

@router.get("/users", response_description="Get all users")
async def get_all_users(token: dict = Depends(decodeJWT)):
    try:
        if not token or token.get("role") != "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Admin access required",
            )
        
        users = await admin_controller.get_all_users()
        return users
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch users: {str(e)}",
        )

@router.get("/users/{user_id}", response_description="Get user profile")
async def get_user(user_id: str, token: dict = Depends(decodeJWT)):
    try:
        if not token or token.get("role") != "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Admin access required",
            )
        
        user = await admin_controller.get_user_profile(user_id)
        return user
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch user: {str(e)}",
        )

@router.put("/users/{user_id}", response_description="Update user details")
async def update_user(user_id: str, update_data: dict = Body(...), token: dict = Depends(decodeJWT)):
    try:
        if not token or token.get("role") != "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Admin access required",
            )
        
        updated_user = await admin_controller.update_user(user_id, update_data, is_admin=True)
        return {
            "message": "User updated successfully",
            "user": updated_user
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update user: {str(e)}",
        )
