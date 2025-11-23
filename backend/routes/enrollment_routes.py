from fastapi import APIRouter, Body, HTTPException, status, Depends
from backend.models.enrollment_model import Enrollment
from backend.utils.jwt_handler import decodeJWT
from backend.database.connection import enrollment_collection, department_collection, user_collection
from datetime import datetime
from bson import ObjectId

router = APIRouter(
    prefix="/enrollments",
    tags=["Enrollments"],
)

@router.get("/societies", response_description="Get all available societies/departments")
async def get_societies(token: dict = Depends(decodeJWT)):
    """Get all departments that students can enroll in (only those with approved teachers)"""
    try:
        if not token:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Authentication required",
            )
        
        user_role = token.get("role")
        
        # Get departments based on role
        if user_role == "student":
            # Students see all departments so they can enroll in any
            departments = await department_collection.find().to_list(1000)
            for dept in departments:
                dept["_id"] = str(dept["_id"])
            
            # Get student's enrollment status for each department
            user_id = token.get("user_id")
            enrollments = await enrollment_collection.find({"user_id": user_id}).to_list(1000)
            enrollment_map = {e["department_id"]: {"status": e["status"], "id": str(e["_id"])} for e in enrollments}
            
            for dept in departments:
                enrollment_info = enrollment_map.get(dept["_id"], {"status": "not_enrolled", "id": None})
                dept["enrollment_status"] = enrollment_info["status"]
                if enrollment_info["id"]:
                    dept["enrollment_id"] = enrollment_info["id"]
        
        elif user_role == "teacher":
            # Teachers see all departments to request enrollment
            departments = await department_collection.find().to_list(1000)
            for dept in departments:
                dept["_id"] = str(dept["_id"])
            
            # Get teacher's enrollment status for all departments
            user_id = token.get("user_id")
            enrollments = await enrollment_collection.find({"user_id": user_id}).to_list(1000)
            enrollment_map = {e["department_id"]: {"status": e["status"], "id": str(e["_id"])} for e in enrollments}
            
            for dept in departments:
                enrollment_info = enrollment_map.get(dept["_id"], {"status": "not_enrolled", "id": None})
                dept["enrollment_status"] = enrollment_info["status"]
                if enrollment_info["id"]:
                    dept["enrollment_id"] = enrollment_info["id"]
        
        else:
            # Admins see all departments
            departments = await department_collection.find().to_list(1000)
            for dept in departments:
                dept["_id"] = str(dept["_id"])
        
        return departments
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] Failed to fetch societies: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch societies: {str(e)}",
        )

@router.post("/request", response_description="Request enrollment to a society")
async def request_enrollment(department_id: str = Body(..., embed=True), token: dict = Depends(decodeJWT)):
    """Student or Teacher requests to enroll in a society/department"""
    try:
        if not token or token.get("role") not in ["student", "teacher"]:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Only students and teachers can request enrollment",
            )
        
        user_id = token.get("user_id")
        user_role = token.get("role")
        
        # Check if already enrolled or pending in THIS specific department
        existing = await enrollment_collection.find_one({
            "user_id": user_id,
            "department_id": department_id
        })
        
        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"You already have a {existing['status']} enrollment request",
            )
        
        # Create enrollment request
        enrollment_data = {
            "user_id": user_id,
            "department_id": department_id,
            "status": "pending",
            "requested_at": datetime.now().isoformat(),
            "reviewed_at": None,
            "reviewed_by": None
        }
        
        result = await enrollment_collection.insert_one(enrollment_data)
        created = await enrollment_collection.find_one({"_id": result.inserted_id})
        
        # Convert ObjectId to string
        created["_id"] = str(created["_id"])
        
        return {
            "message": "Enrollment request submitted successfully",
            "enrollment": created
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to request enrollment: {str(e)}",
        )

@router.get("/my-enrollments", response_description="Get student's enrollments")
async def get_my_enrollments(token: dict = Depends(decodeJWT)):
    """Get current user's enrollment requests"""
    try:
        if not token:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Authentication required",
            )
        
        user_id = token.get("user_id")
        enrollments = await enrollment_collection.find({"user_id": user_id}).to_list(1000)
        
        # Populate department info and convert ObjectId to string
        for enrollment in enrollments:
            enrollment["_id"] = str(enrollment["_id"])
            
            # Handle department_id - it might be string, ObjectId, or None
            dept_id = enrollment.get("department_id")
            if dept_id is not None and dept_id != "" and dept_id != "None":
                # Convert to ObjectId if it's a string
                if isinstance(dept_id, str):
                    dept_id = ObjectId(dept_id)
                dept = await department_collection.find_one({"_id": dept_id})
                enrollment["department_name"] = dept.get("name", "Unknown") if dept else "Unknown"
            else:
                enrollment["department_name"] = "Unknown"
        
        return enrollments
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] Failed to fetch enrollments: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch enrollments: {str(e)}",
        )

@router.get("/pending", response_description="Get pending enrollment requests")
async def get_pending_enrollments(token: dict = Depends(decodeJWT)):
    """Get pending enrollments based on role:
    - Admins see teacher enrollment requests
    - Teachers see student enrollment requests for their department"""
    try:
        if not token or token.get("role") not in ["teacher", "admin"]:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Only teachers and admins can view enrollment requests",
            )
        
        user_role = token.get("role")
        user_id = token.get("user_id")
        
        enrollments = []
        
        if user_role == "admin":
            # Admin sees teacher enrollment requests
            all_pending = await enrollment_collection.find({"status": "pending"}).to_list(1000)
            for enrollment in all_pending:
                user = await user_collection.find_one({"_id": ObjectId(enrollment["user_id"])})
                if user and user.get("role") == "teacher":
                    enrollment["_id"] = str(enrollment["_id"])
                    
                    # Handle department_id - it might be string, ObjectId, or None
                    dept_id = enrollment.get("department_id")
                    if dept_id is not None and dept_id != "" and dept_id != "None":
                        if isinstance(dept_id, str):
                            dept_id = ObjectId(dept_id)
                        dept = await department_collection.find_one({"_id": dept_id})
                        enrollment["department_name"] = dept.get("name", "Unknown") if dept else "Unknown"
                    else:
                        enrollment["department_name"] = "Unknown"
                    
                    enrollment["user_name"] = user.get("name", "Unknown")
                    enrollment["user_email"] = user.get("email", "Unknown")
                    # include the student's registration-entered ID number (fallback to 'N/A')
                    id_num = user.get("id_number")
                    enrollment["user_id_number"] = str(id_num) if id_num else "N/A"
                    enrollment["user_role"] = user.get("role", "Unknown")
                    enrollments.append(enrollment)
        else:
            # Teachers see student enrollment requests for ALL their approved departments
            # Get all teacher's approved departments
            teacher_enrollments = await enrollment_collection.find({
                "user_id": user_id,
                "status": "approved"
            }).to_list(1000)
            
            # Collect valid department IDs
            teacher_dept_ids = []
            for enrollment in teacher_enrollments:
                dept_id = enrollment.get("department_id")
                if dept_id and dept_id != "" and dept_id != "None":
                    teacher_dept_ids.append(dept_id)
            
            if not teacher_dept_ids:
                return []
            
            # Get pending student enrollments for ALL teacher's departments
            all_pending = await enrollment_collection.find({
                "status": "pending",
                "department_id": {"$in": teacher_dept_ids}
            }).to_list(1000)
            
            for enrollment in all_pending:
                user = await user_collection.find_one({"_id": ObjectId(enrollment["user_id"])})
                if user and user.get("role") == "student":
                    enrollment["_id"] = str(enrollment["_id"])
                    
                    # Handle department_id - it might be string, ObjectId, or None
                    dept_id = enrollment.get("department_id")
                    if dept_id is not None and dept_id != "" and dept_id != "None":
                        if isinstance(dept_id, str):
                            dept_id = ObjectId(dept_id)
                        dept = await department_collection.find_one({"_id": dept_id})
                        enrollment["department_name"] = dept.get("name", "Unknown") if dept else "Unknown"
                    else:
                        enrollment["department_name"] = "Unknown"
                    
                    enrollment["user_id"] = str(user["_id"])
                    enrollment["user_name"] = user.get("name", "Unknown")
                    enrollment["user_email"] = user.get("email", "Unknown")
                    # include student's registration id number so teachers can see it when reviewing
                    print(f"DEBUG: User {user.get('email')} - id_number: {user.get('id_number')}")
                    id_num = user.get("id_number")
                    enrollment["user_id_number"] = str(id_num) if id_num else "N/A"
                    enrollment["user_role"] = user.get("role", "Unknown")
                    enrollment["department_id"] = str(enrollment.get("department_id", ""))
                    enrollments.append(enrollment)
        
        return enrollments
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] Failed to fetch pending enrollments: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch pending enrollments: {str(e)}",
        )

@router.get("/approved", response_description="Get approved students")
async def get_approved_enrollments(department_id: str = None, token: dict = Depends(decodeJWT)):
    """Teachers can view all approved students in their department(s)"""
    try:
        if not token or token.get("role") != "teacher":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Only teachers can view approved enrollments",
            )
        
        user_id = token.get("user_id")
        
        # Get teacher's approved departments
        teacher_enrollments = await enrollment_collection.find({
            "user_id": user_id,
            "status": "approved"
        }).to_list(1000)
        
        if not teacher_enrollments:
            return []
        
        # Get list of teacher's approved department IDs
        teacher_dept_ids = [e["department_id"] for e in teacher_enrollments]
        
        # If department_id is specified, verify teacher has access and use it
        # Otherwise, return students from all teacher's departments
        if department_id:
            if department_id not in teacher_dept_ids:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="You don't have access to this department",
                )
            target_dept_ids = [department_id]
        else:
            target_dept_ids = teacher_dept_ids
        
        # Get approved student enrollments for these departments
        approved_enrollments = await enrollment_collection.find({
            "status": "approved",
            "department_id": {"$in": target_dept_ids}
        }).to_list(1000)
        
        result = []
        for enrollment in approved_enrollments:
            user = await user_collection.find_one({"_id": ObjectId(enrollment["user_id"])})
            if user and user.get("role") == "student":
                enrollment["_id"] = str(enrollment["_id"])
                
                # Handle department_id - it might be string, ObjectId, or None
                dept_id = enrollment.get("department_id")
                dept = None
                if dept_id is not None and dept_id != "" and dept_id != "None":
                    if isinstance(dept_id, str):
                        dept_id = ObjectId(dept_id)
                    dept = await department_collection.find_one({"_id": dept_id})
                
                # Build full name from first_name + last_name if available
                first_name = user.get("first_name", "")
                last_name = user.get("last_name", "")
                if first_name and last_name:
                    full_name = f"{first_name} {last_name}"
                else:
                    full_name = user.get("name", "Unknown")
                
                id_num = user.get("id_number")
                user_id_number = str(id_num) if id_num else "N/A"
                
                result.append({
                    "enrollment_id": str(enrollment["_id"]),
                    "user_id": str(user["_id"]),
                    "user_name": full_name,
                    "user_first_name": first_name,
                    "user_last_name": last_name,
                    "user_id_number": user_id_number,
                    "user_email": user.get("email", "Unknown"),
                    "user_role": user.get("role", "Unknown"),
                    "department_name": dept.get("name", "Unknown") if dept else "Unknown",
                    "department_id": enrollment["department_id"],
                    "approved_at": enrollment.get("reviewed_at", enrollment.get("requested_at"))
                })
        
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch approved enrollments: {str(e)}",
        )

@router.get("/approved/export", response_description="Export approved students to Excel")
async def export_approved_enrollments(department_id: str = None, token: dict = Depends(decodeJWT)):
    """Export approved students list to Excel"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    
    try:
        if not token or token.get("role") != "teacher":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Only teachers can export enrollments",
            )
        
        user_id = token.get("user_id")
        
        # Get teacher's approved departments
        teacher_enrollments = await enrollment_collection.find({
            "user_id": user_id,
            "status": "approved"
        }).to_list(1000)
        
        if not teacher_enrollments:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="No approved departments found",
            )
        
        teacher_dept_ids = [e["department_id"] for e in teacher_enrollments]
        
        # Verify access if department_id specified
        if department_id:
            if department_id not in teacher_dept_ids:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="You don't have access to this department",
                )
            target_dept_ids = [department_id]
        else:
            target_dept_ids = teacher_dept_ids
        
        # Get approved students
        approved_enrollments = await enrollment_collection.find({
            "status": "approved",
            "department_id": {"$in": target_dept_ids}
        }).to_list(1000)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Enrolled Students"
        
        # Get department name for header
        dept_name = "All Departments"
        if department_id:
            dept = await department_collection.find_one({"_id": ObjectId(department_id)})
            if dept:
                dept_name = dept.get("name", "Unknown")
        
        # Header
        ws.merge_cells('A1:F1')
        ws['A1'] = f"Enrolled Students - {dept_name}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Column headers
        headers = ['No.', 'ID Number', 'First Name', 'Last Name', 'Email', 'Society', 'Approved Date']
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        row_num = 4
        for idx, enrollment in enumerate(approved_enrollments, 1):
            user = await user_collection.find_one({"_id": ObjectId(enrollment["user_id"])})
            if user and user.get("role") == "student":
                # Handle department_id - it might be string, ObjectId, or None
                dept_id = enrollment.get("department_id")
                dept = None
                if dept_id is not None and dept_id != "" and dept_id != "None":
                    if isinstance(dept_id, str):
                        dept_id = ObjectId(dept_id)
                    dept = await department_collection.find_one({"_id": dept_id})
                
                first_name = user.get("first_name", "")
                last_name = user.get("last_name", "")
                if not first_name or not last_name:
                    full_name = user.get("name", "Unknown")
                    name_parts = full_name.split(" ", 1)
                    first_name = name_parts[0] if not first_name and len(name_parts) > 0 else first_name
                    last_name = name_parts[1] if not last_name and len(name_parts) > 1 else last_name
                
                ws.cell(row=row_num, column=1, value=idx)
                ws.cell(row=row_num, column=2, value=user.get("id_number", "N/A"))
                ws.cell(row=row_num, column=3, value=first_name)
                ws.cell(row=row_num, column=4, value=last_name)
                ws.cell(row=row_num, column=5, value=user.get("email", ""))
                ws.cell(row=row_num, column=6, value=dept.get("name", "Unknown") if dept else "Unknown")
                
                approved_date = enrollment.get("reviewed_at", enrollment.get("requested_at"))
                ws.cell(row=row_num, column=7, value=approved_date)
                
                row_num += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 20
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        filename = f"enrolled_students_{dept_name.replace(' ', '_')}.xlsx"
        
        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to export enrollments: {str(e)}",
        )

@router.get("/teacher/departments", response_description="Get teacher's approved departments")
async def get_teacher_departments(token: dict = Depends(decodeJWT)):
    """Get list of departments teacher is approved in"""
    try:
        if not token or token.get("role") != "teacher":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Only teachers can access this endpoint",
            )
        
        user_id = token.get("user_id")
        
        # Get teacher's approved departments
        teacher_enrollments = await enrollment_collection.find({
            "user_id": user_id,
            "status": "approved"
        }).to_list(1000)
        
        departments = []
        for enrollment in teacher_enrollments:
            # Handle department_id - it might be string, ObjectId, or None
            dept_id = enrollment.get("department_id")
            dept = None
            if dept_id is not None and dept_id != "" and dept_id != "None":
                if isinstance(dept_id, str):
                    dept_id = ObjectId(dept_id)
                dept = await department_collection.find_one({"_id": dept_id})
            
            if dept:
                departments.append({
                    "department_id": str(dept["_id"]),
                    "department_name": dept.get("name", "Unknown"),
                    "enrollment_id": str(enrollment["_id"])
                })
        
        return departments
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to fetch teacher departments: {str(e)}",
        )

@router.post("/review/{enrollment_id}", response_description="Approve or decline enrollment")
async def review_enrollment(
    enrollment_id: str, 
    action: str = Body(..., embed=True),
    token: dict = Depends(decodeJWT)
):
    """Teacher approves or declines an enrollment request"""
    try:
        if not token or token.get("role") not in ["teacher", "admin"]:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Only teachers and admins can review enrollments",
            )
        
        if action not in ["approved", "declined"]:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Action must be 'approved' or 'declined'",
            )
        
        # Update enrollment
        result = await enrollment_collection.update_one(
            {"_id": ObjectId(enrollment_id)},
            {
                "$set": {
                    "status": action,
                    "reviewed_at": datetime.now().isoformat(),
                    "reviewed_by": token.get("user_id")
                }
            }
        )
        
        if result.modified_count == 0:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Enrollment not found",
            )
        
        return {
            "message": f"Enrollment {action} successfully",
            "status": action
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to review enrollment: {str(e)}",
        )

@router.delete("/{enrollment_id}", response_description="Cancel/Leave enrollment")
async def cancel_enrollment(enrollment_id: str, token: dict = Depends(decodeJWT)):
    """Allow teachers to cancel their enrollment in a society"""
    try:
        if not token:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Authentication required",
            )
        
        user_id = token.get("user_id")
        user_role = token.get("role")
        
        # Get the enrollment
        enrollment = await enrollment_collection.find_one({"_id": ObjectId(enrollment_id)})
        
        if not enrollment:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Enrollment not found",
            )
        
        # Check if the user owns this enrollment
        if enrollment["user_id"] != user_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only cancel your own enrollments",
            )
        
        # Teachers can cancel their approved enrollments
        if user_role == "teacher" and enrollment["status"] == "approved":
            # Delete the enrollment
            result = await enrollment_collection.delete_one({"_id": ObjectId(enrollment_id)})
            
            if result.deleted_count == 0:
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail="Failed to cancel enrollment",
                )
            
            return {
                "message": "Successfully left the society",
                "enrollment_id": enrollment_id
            }
        elif enrollment["status"] == "pending":
            # Anyone can cancel pending enrollments
            result = await enrollment_collection.delete_one({"_id": ObjectId(enrollment_id)})
            
            if result.deleted_count == 0:
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail="Failed to cancel enrollment",
                )
            
            return {
                "message": "Enrollment request cancelled",
                "enrollment_id": enrollment_id
            }
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Cannot cancel this enrollment",
            )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to cancel enrollment: {str(e)}",
        )
